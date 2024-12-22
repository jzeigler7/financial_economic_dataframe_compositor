import openpyxl
from openpyxl import load_workbook
import os

def delete_specific_columns(input_file, output_file, columns_to_delete):
    """
    Deletes specified columns from an Excel file.

    :param input_file: Path to the input Excel file.
    :param output_file: Path to save the cleaned Excel file.
    :param columns_to_delete: A list of column letters to delete (e.g., ['IH', 'AGY']).
    """
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"The file '{input_file}' does not exist in the current directory.")

    # Load the workbook and select the active worksheet
    wb = load_workbook(input_file)
    ws = wb.active

    # Convert column letters to indices and sort them in reverse order
    # Sorting in reverse ensures that deleting one column doesn't affect the indices of the remaining columns
    columns_indices = sorted(
        [openpyxl.utils.column_index_from_string(col) for col in columns_to_delete],
        reverse=True
    )

    for col_idx in columns_indices:
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        print(f"Deleting column {col_letter}...")
        ws.delete_cols(col_idx, 1)

    # Save the modified workbook
    wb.save(output_file)
    print(f"Specified columns have been deleted and saved to '{output_file}' successfully.")

def main():
    input_filename = 'batch1.xlsx'
    output_filename = 'batch1_cleaned.xlsx'
    columns_to_remove = ['IH', 'AGY', 'GYV', 'HBS']

    try:
        delete_specific_columns(
            input_file=input_filename,
            output_file=output_filename,
            columns_to_delete=columns_to_remove
        )
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    main()
