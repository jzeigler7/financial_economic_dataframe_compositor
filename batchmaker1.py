import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import copy

def concatenate_excels(file_names, output_file):
    # Load all workbooks and their active sheets
    workbooks = []
    sheets = []
    for file in file_names:
        if not os.path.exists(file):
            raise FileNotFoundError(f"The file {file} does not exist in the current directory.")
        wb = openpyxl.load_workbook(file, data_only=False)
        ws = wb.active
        workbooks.append(wb)
        sheets.append(ws)

    # Determine the maximum number of rows among all sheets
    max_rows = max(ws.max_row for ws in sheets)

    # Determine the maximum number of columns in each sheet
    max_columns = [ws.max_column for ws in sheets]

    # Create a new workbook and select the active worksheet
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "Batch1"

    # Initialize the starting column for concatenation
    current_col = 1

    for idx, ws in enumerate(sheets):
        print(f"Processing sheet from {file_names[idx]}...")
        # Iterate through each row in the current sheet
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=current_col + cell.column - 1, value=cell.value)

                # Copy cell style if it has a style
                if cell.has_style:
                    new_cell.font = copy.copy(cell.font)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy.copy(cell.protection)
                    new_cell.alignment = copy.copy(cell.alignment)

        # Update the current_col to the next available column
        current_col += ws.max_column

    # Adjust column widths (optional but recommended)
    for column_cells in new_ws.columns:
        max_length = 0
        column = column_cells[0].column  # Get the column name (A, B, C, ...)
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        adjusted_width = (max_length + 2)
        new_ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    # Save the concatenated workbook
    new_wb.save(output_file)
    print(f"All sheets have been concatenated into {output_file} successfully.")

def main():
    # List of source Excel files to concatenate
    source_files = [
        'market_data.xlsx',
        'market_stats.xlsx',
        'ta_data.xlsx',
        'economic_data.xlsx',
        'economic_stats.xlsx'
    ]

    output_filename = 'batch1.xlsx'

    try:
        concatenate_excels(source_files, output_filename)
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    main()
