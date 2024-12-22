import openpyxl
from openpyxl.styles import PatternFill
import multiprocessing
import os
import sys

def color_spreadsheet(file_color):
    """
    Colors all cells in the given Excel file with the specified color.

    :param file_color: A tuple containing the filename and the HEX color code.
    """
    filename, color_code = file_color
    try:
        # Define the fill pattern with the specified color
        fill = PatternFill(start_color=color_code, end_color=color_code, fill_type='solid')

        # Load the workbook
        wb = openpyxl.load_workbook(filename)

        # Iterate through all worksheets
        for sheet in wb.worksheets:
            # Iterate through all rows and cells
            for row in sheet.iter_rows():
                for cell in row:
                    cell.fill = fill

        # Save the workbook
        wb.save(filename)
        print(f"Successfully colored '{filename}' with color #{color_code}.")
    except Exception as e:
        print(f"Error processing '{filename}': {e}")

def main():
    # Define the directory where the Excel files are located (same as the script's directory)
    script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))

    # List of Excel files and their corresponding color codes
    files_and_colors = [
        ('market_data.xlsx', 'F08080'),     # Light Red (Coral)
        ('market_stats.xlsx', 'FFFFE0'),    # Light Yellow
        ('ta_data.xlsx', 'ADD8E6'),         # Light Blue
        ('economic_data.xlsx', 'E6E6FA'),   # Light Purple
    ]

    # Prepend the script directory to the filenames to ensure correct paths
    files_and_colors = [(os.path.join(script_dir, fname), color) for fname, color in files_and_colors]

    # Check if all files exist before proceeding
    missing_files = [fname for fname, _ in files_and_colors if not os.path.isfile(fname)]
    if missing_files:
        print("The following Excel files were not found in the script directory:")
        for mf in missing_files:
            print(f" - {os.path.basename(mf)}")
        sys.exit(1)

    # Create a pool of worker processes equal to the number of CPU cores
    pool = multiprocessing.Pool(processes=multiprocessing.cpu_count())

    try:
        # Map the function to the list of files and colors
        pool.map(color_spreadsheet, files_and_colors)
    finally:
        # Close the pool and wait for the work to finish
        pool.close()
        pool.join()

if __name__ == "__main__":
    main()
