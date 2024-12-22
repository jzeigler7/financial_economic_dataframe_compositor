import os
from openpyxl import load_workbook

def read_missing_count(count_file_path):
    """
    Reads the missing_cells_count.txt file and returns the count as an integer.
    """
    try:
        with open(count_file_path, 'r') as file:
            content = file.read().strip()
            missing_count = int(content)
            if missing_count < 0:
                raise ValueError("The number of missing cells cannot be negative.")
            return missing_count
    except FileNotFoundError:
        print(f"Error: '{count_file_path}' not found.")
        return None
    except ValueError:
        print(f"Error: The file '{count_file_path}' must contain a non-negative integer.")
        return None

def delete_rows_in_workbook(file_path, start_row, end_row):
    """
    Deletes rows from start_row to end_row (inclusive) in the given Excel workbook.
    """
    try:
        wb = load_workbook(filename=file_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Delete from end_row to start_row to avoid shifting issues
            for row in range(end_row, start_row - 1, -1):
                if row <= ws.max_row:
                    ws.delete_rows(row)
                else:
                    print(f"Warning: Sheet '{sheet_name}' in '{file_path}' does not have row {row}.")
        wb.save(file_path)
        print(f"Successfully updated '{file_path}'. Deleted rows {start_row} through {end_row}.")
    except FileNotFoundError:
        print(f"Error: '{file_path}' not found.")
    except Exception as e:
        print(f"An error occurred while processing '{file_path}': {e}")

def main():
    # Define file names
    spreadsheet_files = ['market_data.xlsx', 'market_stats.xlsx', 'ta_data.xlsx']
    count_file = 'missing_cells_count.txt'

    # Get the directory where the script is located
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Path to the count file
    count_path = os.path.join(script_dir, count_file)

    # Read the missing_cells_count.txt
    missing_count = read_missing_count(count_path)
    if missing_count is None:
        return

    # Define the range of rows to delete (1-based indexing for Excel)
    start_row = 2
    end_row = 2 + missing_count  # Inclusive

    # Process each spreadsheet
    for file_name in spreadsheet_files:
        file_path = os.path.join(script_dir, file_name)
        delete_rows_in_workbook(file_path, start_row, end_row)

if __name__ == "__main__":
    main()


