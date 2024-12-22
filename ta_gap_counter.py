from openpyxl import load_workbook

# File paths
input_file = r'C:\Users\jacob\OneDrive\Documents\ZAP\ta_data.xlsx'
output_file = r'C:\Users\jacob\OneDrive\Documents\ZAP\missing_cells_count.txt'

# Load the Excel workbook
wb = load_workbook(input_file, data_only=True)
ws = wb.active

# Start counting from row 4
missing_rows_count = 0
found_full_row = False

# Iterate through each row starting from row 4
for row_idx, row in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row), start=4):
    for col_idx, cell in enumerate(row, start=1):
        if cell.value is None:
            # Log the first empty cell's coordinates
            print(f"Empty cell found at row {row_idx}, column {col_idx}")
            missing_rows_count += 1
            break  # Move to the next row after finding the first empty cell
    else:
        # Stop if a fully populated row is found
        found_full_row = True
        print(f"Fully populated row found at row {row_idx}. Stopping iteration.")
        break

missing_rows_count = missing_rows_count + 1

# Save the count of rows with missing cells to a .txt file
with open(output_file, 'w') as file:
    file.write(str(missing_rows_count))

print(f"Number of rows with missing cells: {missing_rows_count}")
print(f"Result saved to {output_file}")

