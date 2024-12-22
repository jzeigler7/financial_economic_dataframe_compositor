import yfinance as yf
import pandas as pd
import json
from datetime import datetime
import argparse
import os
from tqdm import tqdm  # For progress bars
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import time

# Setup argument parser to accept a custom day argument
parser = argparse.ArgumentParser(description='Download stock data for tickers and save it to an Excel file.')
parser.add_argument('--day', type=str, help='Specify the current calendar day in YYYY-MM-DD format (optional)')
args = parser.parse_args()

# Load the JSON file with ticker symbols
with open('tickers.json', 'r') as file:
    tickers_data = json.load(file)

# Extract tickers from the market_data section
tickers = list(tickers_data["market_data"].keys())

# Set the start date
start_date = "2007-08-01"

# Determine the current "day"
if args.day:
    try:
        end_date = datetime.strptime(args.day, '%Y-%m-%d').strftime('%Y-%m-%d')
    except ValueError:
        print("Error: Invalid date format. Please use YYYY-MM-DD.")
        exit(1)
else:
    end_date = datetime.today().strftime('%Y-%m-%d')

# Check if the Excel file already exists
output_file = 'market_data.xlsx'
existing_data = None

if os.path.exists(output_file):
    print(f"Found existing file: {output_file}. Loading it to check for gaps.")

    # Load the existing spreadsheet
    existing_data = pd.read_excel(output_file)

    # Find the most recent date in the existing file
    last_recorded_date = existing_data['Date'].max()

    # Set the start date for new data to be the day after the last recorded date
    start_date = (pd.to_datetime(last_recorded_date) + pd.DateOffset(1)).strftime('%Y-%m-%d')

    print(f"Fetching data starting from {start_date} to fill the gap.")

# If there's no gap or nothing to append, exit the program
if start_date >= end_date:
    print(f"Data is already up-to-date. No new data to fetch.")
    exit(0)

# Create a dictionary to hold all ticker data
all_ticker_data = {}

# Loop through all tickers and download the historical market data for the gap period
print("\nDownloading ticker data:")
for ticker in tqdm(tickers):
    while True:
        try:
            data = yf.download(ticker, start=start_date, end=end_date, progress=False)

            # Ensure we only keep the desired columns (Open, High, Low, Close, Adj Close, Volume)
            if not data.empty:
                all_ticker_data[ticker] = data[['Open', 'High', 'Low', 'Close', 'Adj Close', 'Volume']]
                break  # Exit the retry loop on successful download
            else:
                print(f"No data found for {ticker}. Retrying in 10 seconds...")
                time.sleep(10)  # Wait before retrying
        except Exception as e:
            print(f"Error downloading {ticker}: {e}. Retrying in 10 seconds...")
            time.sleep(10)  # Wait before retrying

# Concatenate all the data into one DataFrame with multi-level columns
if all_ticker_data:
    all_dates = pd.Index(sorted(set().union(*(data.index for data in all_ticker_data.values()))))
    for ticker in all_ticker_data:
        all_ticker_data[ticker] = all_ticker_data[ticker].reindex(all_dates)

    new_data_df = pd.concat(all_ticker_data.values(), keys=all_ticker_data.keys(), axis=1)
    new_data_df.columns = ['_'.join(col).strip() for col in new_data_df.columns.values]
    new_data_df.reset_index(inplace=True)
    new_data_df.rename(columns={'index': 'Date'}, inplace=True)

    # Apply row filtering, interpolation, etc.
    data_columns = new_data_df.columns.drop('Date')
    fraction_missing = new_data_df[data_columns].isnull().mean(axis=1)
    num_rows_before = new_data_df.shape[0]
    new_data_df = new_data_df[fraction_missing < 0.25].reset_index(drop=True)
    num_rows_after = new_data_df.shape[0]

    print(f"\nNew data: Rows before pruning: {num_rows_before}, Rows after pruning: {num_rows_after}")

    new_data_df.set_index('Date', inplace=True)
    new_data_df.index = pd.to_datetime(new_data_df.index)
    new_data_df = new_data_df.interpolate(method='time', limit_direction='both')
    new_data_df = new_data_df.fillna(method='ffill').fillna(method='bfill')
    new_data_df.reset_index(inplace=True)
    empty_cols_new = new_data_df.columns[new_data_df.isnull().all()]
    if not empty_cols_new.empty:
        print(f"Removing empty columns from new data: {list(empty_cols_new)}")
        new_data_df.drop(columns=empty_cols_new, inplace=True)

    new_data_df.dropna(how='all', inplace=True)

    # Combine existing data and new data if necessary
    if existing_data is not None:
        combined_data = pd.concat([existing_data, new_data_df], ignore_index=True)
        combined_data.drop_duplicates(subset='Date', keep='last', inplace=True)
    else:
        combined_data = new_data_df

    empty_cols_combined = combined_data.columns[combined_data.isnull().all()]
    if not empty_cols_combined.empty:
        print(f"Removing empty columns from combined data: {list(empty_cols_combined)}")
        combined_data.drop(columns=empty_cols_combined, inplace=True)

    combined_data.dropna(how='all', inplace=True)
    combined_data.dropna(subset=['Date'], inplace=True)
    combined_data['Date'] = pd.to_datetime(combined_data['Date'])
    combined_data.reset_index(drop=True, inplace=True)

    # Write the data to Excel without coloring
    wb = Workbook()
    ws = wb.active

    # Get the exact dimensions of the DataFrame
    num_rows, num_cols = combined_data.shape

    # Write the data to Excel without coloring
    for r_idx, row in enumerate(dataframe_to_rows(combined_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            if c_idx <= num_cols:  # Only write as many columns as the data requires
                ws.cell(row=r_idx, column=c_idx, value=value)

    # Adjust the column width for better visibility
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # **Remove Time from Date Column**
    # Assuming 'Date' is the first column (Column A)
    date_column = 'A'
    date_format = 'YYYY-MM-DD'  # Excel date format without time

    # Iterate over all cells in the Date column (starting from row 2 to skip header)
    for cell in ws[date_column][1:]:  # ws['A'][0] is the header
        if isinstance(cell.value, datetime):
            cell.number_format = date_format
            # Optionally, ensure only date is present
            cell.value = cell.value.date()

    # Save the file
    wb.save(output_file)
    print(f"\nNew data appended and saved to {output_file}")
else:
    print("No new data to append.")