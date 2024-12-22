import openpyxl
import csv
from datetime import datetime

# Load the CSV file and extract the dates
csv_file_path = 'found_dates.csv'
print("Loading CSV file...")

with open(csv_file_path, 'r') as csvfile:
    reader = csv.reader(csvfile)
    next(reader)  # Skip header
    csv_dates = {row[0] for row in reader}

print(f"CSV dates loaded: {list(csv_dates)[:5]}... (first 5)")

# Function to process an Excel file and delete rows absent from CSV
def process_excel_file(excel_file_path):
    print(f"Loading Excel file: {excel_file_path}...")
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active
    print(f"Excel file {excel_file_path} loaded.")

    # Iterate through the Excel file rows and find rows where the date is absent from CSV
    rows_to_delete = []
    for row in range(2, sheet.max_row + 1):  # Assuming first row is a header
        excel_date_cell = sheet.cell(row=row, column=1).value
        if isinstance(excel_date_cell, datetime):
            excel_date_str = excel_date_cell.strftime("%Y-%m-%d")
        else:
            excel_date_str = str(excel_date_cell)

        if excel_date_str not in csv_dates:
            rows_to_delete.append(row)

    print(f"Rows to be deleted in {excel_file_path}: {rows_to_delete}")

    # Delete rows from the bottom to top to avoid shifting issues
    for row in reversed(rows_to_delete):
        sheet.delete_rows(row)

    # Save the modified Excel file
    workbook.save(excel_file_path)
    print(f"Rows deleted and Excel file {excel_file_path} saved.")

# Process both Excel files
process_excel_file('economic_data.xlsx')
process_excel_file('economic_stats.xlsx')

